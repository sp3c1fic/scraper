import pandas as pd
import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from datetime import datetime
from reader import DataReader

INITIAL_COL_OFFSET = 3
COL_NAME = 'Expected delivery'

def iterate_dataframes(update_progress, name, output_dir):
    
        excel_file_path = 'Modified_plantool.xlsx' # no hardcoding // this reads the example 'template' file also provides the first dataframe move it to config.json
        sheet_name = 'Sheet1' # no hardcoding // the sheet name where the second data frame info is taken from

        inventory_df = pd.read_excel(excel_file_path, sheet_name=sheet_name) # creating the inventory_sheet dataframe
        store_data_df = pd.read_excel(name) # this should not be hardcoded // the scraped stock info dataframe 
        update_progress(12)
        store_delivery_dates_df = store_data_df['Store Delivery Date']
        update_progress(17)
        first_unnamed = inventory_df['Unnamed: 0']
        update_progress(25)
        second_unnamed = inventory_df['Unnamed: 1']
        update_progress(48)
        fill_in_dates(store_data_df, store_delivery_dates_df, inventory_df, first_unnamed, second_unnamed)
        update_progress(54)
        fill_in_last_column(inventory_df, store_data_df, store_delivery_dates_df, second_unnamed)
        update_progress(74)
        output_dir = define_output_dir(output_dir)
        file_name = f'{output_dir}test_plantool_{get_current_date_time()}.xlsx'

        DataReader.fill_in_last_product(inventory_df, second_unnamed, store_data_df)

        update_progress(90)

        inventory_df = inventory_df.apply(pd.to_numeric, errors='ignore')        
        inventory_df.to_excel(file_name, engine="openpyxl")

        update_progress(95)
        modify_columns(file_name, COL_NAME)
        update_progress(100)


def format_cells(col, header_value, expected_keyword):
      
        if header_value and expected_keyword.lower() in str(header_value).lower():
            for cell in col:
                # Apply the number format to the entire column if the header matches
                if isinstance(cell.value, (int, float)):
                    cell.number_format = u'#,##0.00'

                elif isinstance(cell.value, str) and cell.value.count('-') == 2:
                      try:
                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
                      except ValueError:
                            pass


def modify_columns(file_name, expected_keyword):
    workbook = openpyxl.load_workbook(file_name)
    ws = workbook.active
    header_row = 1
    # Manually set column widths (example: 20 for all columns)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter
        header_value = ws[f"{col_letter}{header_row}"].value
        
        format_cells(col, header_value, expected_keyword)

        for cell in col: # THIS ALSO NEEDS TO BE A SEPARATE FUNCTION
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))      
            except:
                pass
        # Set a width that fits the longest content in the column, or manually set it
        ws.column_dimensions[col_letter].width = max(max_length + 2, 20)
        
    # Save the modified workbook
    workbook.save(file_name)


def define_output_dir(output_dir):
        if output_dir != "" and output_dir is not None:
                return f"{output_dir}/"
        return ""

def is_not_expected_delivery_column(col):
        return f'{COL_NAME}.' not in col and COL_NAME != col

def get_current_date_time():
       return datetime.now().strftime('%d_%m_%Y_%H_%M_%S')

def get_df_columns_len(inventory_df):
      return len(inventory_df.columns)


def get_last_column(inventory_df, col_name):
      return inventory_df.columns.get_loc(col_name)

def add_columns(inventory_df, additional_columns_needed, index, current_date):
        for _ in range(additional_columns_needed):
                inventory_df[f'{COL_NAME}.{index}'] = pd.NA
                inventory_df[f'{COL_NAME}.{index}'][2] = current_date


def fill_in_dates(store_data_df, store_delivery_dates_df, inventory_df, first_unnamed, second_unnamed):
        initial_date = store_delivery_dates_df[0]
        index = 1
        
        for current_date in store_delivery_dates_df.iloc[1:].drop_duplicates():
                if initial_date == current_date:
                        break
                current_date_index = 1
                for col in inventory_df:

                        if current_date_index > len(store_delivery_dates_df):
                            return

                        if is_not_expected_delivery_column(col):
                                continue
                        if f'{COL_NAME}.' in col: # Expected delivery.{index} is the same as the name of the current column
                                last_exp_del_col_ind = get_last_column(inventory_df, col) # the index of the last Expected.delivery column Initially equal to 6
                                insert_pos = last_exp_del_col_ind + INITIAL_COL_OFFSET# new insert position will be index 9 supposedly
                                df_columns_len = get_df_columns_len(inventory_df)
                                if insert_pos >= df_columns_len:                       ## needs to be separated out in an additional function
                                        additional_columns_needed = insert_pos - len(inventory_df.columns) + 1
                                        add_columns(inventory_df, additional_columns_needed, index, current_date)                    
                                DataReader.fill_in_total_values(inventory_df, store_data_df, second_unnamed, col, current_date_index)
                                current_date_index += 1

                        elif COL_NAME == col:
                               inventory_df[COL_NAME][2] = initial_date
                               DataReader.fill_in_total_values(inventory_df, store_data_df, second_unnamed, col, 0)
                index += 1


def fill_in_last_column(inventory_df, store_data_df, store_delivery_dates_df, second_unnamed):
        last_col = inventory_df.iloc[:,-1:]
        last_col_name = list(last_col.columns)[0]
        last_date_index = store_delivery_dates_df.iloc[1:].drop_duplicates().index[-2]
        DataReader.fill_in_total_values(inventory_df, store_data_df, second_unnamed, last_col_name, last_date_index)

